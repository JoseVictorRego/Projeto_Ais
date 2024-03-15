import os
import queue
from openpyxl import load_workbook
from bd_funcao import *

# Obtém o caminho atual do diretório do script_ em execução
caminho_atual = os.path.dirname(os.path.realpath(__file__))

# Junta o caminho atual com o caminho relativo da pasta
pasta = os.path.join(caminho_atual, "arquivo tipo 2")

# Lista todos os arquivos na pasta
arquivos_na_pasta = os.listdir(pasta)

# Criar uma fila e adicionar os arquivos a ela
fila_arquivos = queue.Queue()

for arquivo in arquivos_na_pasta:
    fila_arquivos.put(arquivo)

# Tamanho da lista
total = len(arquivos_na_pasta)
Cont = 1

# Corre os arquivos da fila
while not fila_arquivos.empty():

    # Imprimir nome do arquivo e porcentagem de leitura
    nome_arquivo = fila_arquivos.get()
    print("=> ", nome_arquivo, "-", Cont, "/", total, " <=\n")

    # Junta o caminho atual com o caminho relativo da pasta e do arquivo
    caminho_completo = os.path.join(pasta, nome_arquivo)

    # Carrega a planilha usando o caminho completo
    planilha = load_workbook(caminho_completo)

    aba_ativa = planilha.active

    # Capitura o titulo da primeir coluna do excel para controle das paginas
    Titulo = aba_ativa.cell(row=1, column=1)  # Coluna "A1"

    # inciar coneção com o banco de dados
    conn, cursor = conectar()
    if conn and cursor:
        try:
            

            # inciar coneção com o banco de dados
            for celula in aba_ativa["E"]:

                # Encontrar celula com essa denominação
                if celula.value == "NOME DO (A) ALUNO (A)":

                    # Pular uma linha para poder caputurar os dados dos alunos
                    proxima_linha = celula.row + 1

                    # Caputurar os dados do aluno e salvar nas variaveis em amostra
                    NomeAluno = aba_ativa.cell(row=proxima_linha, column=5)  # Coluna "E"
                    SexoAluno = aba_ativa.cell(row=proxima_linha, column=16)  # Coluna "P"
                    FormIngres = aba_ativa.cell(row=proxima_linha, column=18)  # Coluna "R"
                    PeriodoIngres = aba_ativa.cell(row=proxima_linha, column=28)  # Coluna "AB"

                    if NomeAluno != "DOCENTE(S): -----------.":
                        # Adicionar aluno e recolher seu id_aluno, ou so recolher id aluno
                        id_aluno = consultar_id_aluno(conn, cursor, NomeAluno.value, SexoAluno.value,
                                                      FormIngres.value, PeriodoIngres.value)

                    # Linha do inicío do recolhimento dos dados academicos
                    linhaD_campus = proxima_linha + 5

                    # Iniciar leitura dos dados do historico das disciplinas cursadas
                    for celulaD in aba_ativa["A"][linhaD_campus:]:

                        # if caso a celula seja nula ou não importante para o recolhimento do dados
                        if (celulaD.value is None or celulaD.value == "COMPONENTES CURRICULARES OBRIGATÓRIOS"
                                or celulaD.value == "ANO" or celulaD.value == "ISOLADAS"):
                            continue

                        # Condição para caso dados de disciplinas de eletiva
                        elif (celulaD.value == "CARGA HORÁRIA" or celulaD.value == "TOTAL" or
                              celulaD.value == "OBRIGATORIA"
                              or celulaD.value == "ELETIVA" or celulaD.value == "COMPONENTES ELETIVOS"):
                            continue

                        # Condição para encontrar celula inicial para pular de folha
                        elif celulaD.value == Titulo.value or NomeAluno.value == "DOCENTE(S): -----------.":
                            break

                        # Condição principal para recolher dados da disciplina
                        elif celulaD.value >= PeriodoIngres.value or celulaD.value <= PeriodoIngres.value:

                            Ano = aba_ativa.cell(row=celulaD.row, column=1)  # Coluna "A"
                            Semetre = aba_ativa.cell(row=celulaD.row, column=2)  # Coluna "B"
                            Codigo = aba_ativa.cell(row=celulaD.row, column=3)  # Coluna "C"
                            Disciplina = aba_ativa.cell(row=celulaD.row, column=6)  # Coluna "F"
                            CargaHoraria = aba_ativa.cell(row=celulaD.row, column=19)  # Coluna "S"
                            Creditos = aba_ativa.cell(row=celulaD.row, column=21)  # Coluna "U"
                            NotaFinal = aba_ativa.cell(row=celulaD.row, column=24)  # Coluna "X"
                            Situacao = aba_ativa.cell(row=celulaD.row, column=26)  # Coluna "Z"

                            if NomeAluno != "DOCENTE(S): -----------.":
                                # Adicionar aluno e recolher seu id_aluno, ou so recolher id aluno
                                id_aluno = consultar_id_aluno(conn, cursor, NomeAluno.value, SexoAluno.value,
                                                              FormIngres.value, PeriodoIngres.value)

                            # criar periodoLetivo
                            periodoLetivo = Ano.value + Semetre.value * 0.1

                            # Separar Disciplinas em nome_disciplina e nome_professor
                            if '\n' in Disciplina.value:

                                # Distribuir os dados para as variaveis
                                nome_disciplina = Disciplina.value.split('\n')[0]
                                nome_professor = Disciplina.value.split('\nDOCENTE(S): ')[1]

                                # Identificar id_professor ou adicionar no banco de dados
                                id_professor = consultar_id_professor(conn, cursor, nome_professor)

                            else:
                                # Distribuir os dados para as variaveis
                                nome_disciplina = Disciplina.value
                                id_professor = None

                            # Indentificar id_disciplina ou adicionar no banco de dados
                            id_disciplina = consultar_id_disciplina(conn, cursor, Codigo.value, nome_disciplina,
                                                                    CargaHoraria.value, Creditos.value)

                            # Adiscipnar dados do historico na base de dados ou informar se ele ja existir nela
                            inserir_dados_historico(conn, cursor, periodoLetivo, id_aluno, id_disciplina,
                                                    id_professor, NotaFinal.value, Situacao.value)

        # Finalizar conecção com o banco de dados
        finally:
            desconectar(conn, cursor)
    # Contador de arquivos lidos
    Cont = Cont + 1
    print("=============================================================================================")